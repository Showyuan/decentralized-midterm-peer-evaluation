#!/usr/bin/env python3
"""
結果收集器 (Web 版本)
從資料庫收集評分結果並輸出為 JSON 和 Excel 格式
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List
from collections import defaultdict

from stage2_db_manager import DatabaseManager

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  警告: openpyxl 未安裝，無法輸出 Excel 檔案")


class ResultCollectorWeb:
    """從資料庫收集評分結果"""
    
    def __init__(self):
        self.db = DatabaseManager()
        
        # 從配置獲取路徑
        try:
            from stage0_config_unified import PeerEvaluationConfig
        except ImportError:
            from .stage0_config_unified import PeerEvaluationConfig
        
        self.config = PeerEvaluationConfig()
        self.output_dir = Path(self.config.ensure_output_dir('stage5_results'))
        
        # 載入學生資料
        self.students = self._load_student_data()
    
    def _load_student_data(self) -> Dict:
        """載入學生資料"""
        data_path = Path(self.config.get_path('stage1_output', 'processed_data'))
        
        if data_path.exists():
            with open(data_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('students', {})
        
        return {}
    
    def collect_results(self) -> Dict:
        """
        收集所有評分結果
        
        Returns:
            Dict: 包含所有評分結果的字典
        """
        print("=" * 80)
        print("📊 開始從資料庫收集評分結果")
        print("=" * 80)
        
        # 獲取所有提交記錄
        submissions = self.db.get_all_submissions()
        
        print(f"📝 總提交記錄數: {len(submissions)}")
        
        # 按被評分者組織數據
        results_by_target = defaultdict(lambda: {
            'evaluations': [],
            'questions': defaultdict(list)
        })
        
        for submission in submissions:
            target_id = submission['target_id']
            evaluator_id = submission['evaluator_id']
            question_id = submission['question_id']
            score = submission['score']
            comment = submission['comment']
            
            # 記錄評分
            results_by_target[target_id]['evaluations'].append({
                'evaluator_id': evaluator_id,
                'question_id': question_id,
                'score': score,
                'comment': comment,
                'submitted_at': submission['submitted_at']
            })
            
            # 按題目分組
            results_by_target[target_id]['questions'][question_id].append({
                'evaluator_id': evaluator_id,
                'score': score,
                'comment': comment
            })
        
        # 統計每個學生的結果
        final_results = {}
        
        for target_id, data in results_by_target.items():
            # 獲取學生資訊
            student_info = self.students.get(target_id, {})
            student_name = student_info.get('name', target_id)
            
            # 計算每題的統計
            question_stats = {}
            total_scores = []
            
            for question_id in sorted(data['questions'].keys()):
                scores = [item['score'] for item in data['questions'][question_id]]
                comments = [item['comment'] for item in data['questions'][question_id] if item['comment']]
                
                question_stats[question_id] = {
                    'scores': scores,
                    'mean': round(sum(scores) / len(scores), 2),
                    'min': min(scores),
                    'max': max(scores),
                    'count': len(scores),
                    'comments': comments
                }
                
                total_scores.extend(scores)
            
            # 計算總分統計
            final_results[target_id] = {
                'student_id': target_id,
                'student_name': student_name,
                'total_evaluations': len(set(e['evaluator_id'] for e in data['evaluations'])),
                'total_scores': len(total_scores),
                'overall_mean': round(sum(total_scores) / len(total_scores), 2) if total_scores else 0,
                'overall_min': min(total_scores) if total_scores else 0,
                'overall_max': max(total_scores) if total_scores else 0,
                'question_stats': question_stats,
                'all_evaluations': data['evaluations']
            }
        
        # 生成匯總統計
        summary = self._generate_summary(final_results)
        
        result = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'total_students': len(final_results),
                'total_submissions': len(submissions),
                'total_evaluations': len(submissions) // 5,  # 每份評分有 5 題
            },
            'summary': summary,
            'results': final_results
        }
        
        return result
    
    def _generate_summary(self, results: Dict) -> Dict:
        """生成匯總統計"""
        
        all_means = [data['overall_mean'] for data in results.values()]
        
        # 按問題統計
        question_summary = defaultdict(list)
        for data in results.values():
            for q_id, stats in data['question_stats'].items():
                question_summary[q_id].append(stats['mean'])
        
        summary = {
            'overall': {
                'mean': round(sum(all_means) / len(all_means), 2) if all_means else 0,
                'min': round(min(all_means), 2) if all_means else 0,
                'max': round(max(all_means), 2) if all_means else 0,
            },
            'by_question': {}
        }
        
        for q_id, means in question_summary.items():
            summary['by_question'][q_id] = {
                'mean': round(sum(means) / len(means), 2),
                'min': round(min(means), 2),
                'max': round(max(means), 2),
            }
        
        return summary
    
    def export_to_json(self, results: Dict, filename: str = None) -> Path:
        """
        匯出結果為 JSON 檔案
        
        Args:
            results: 結果字典
            filename: 檔案名稱（可選）
            
        Returns:
            Path: 輸出檔案路徑
        """
        if filename is None:
            # 使用配置生成檔案路徑
            output_path = Path(self.config.get_path('stage5_results', 'evaluation_results_json', timestamp=datetime.now()))
        else:
            output_path = self.output_dir / filename
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        print(f"\n✅ 結果已匯出: {output_path}")
        print(f"   檔案大小: {output_path.stat().st_size / 1024:.2f} KB")
        
        return output_path
    
    def export_for_vancouver(self, results: Dict) -> Path:
        """
        匯出為 Vancouver 系統相容格式
        
        Vancouver 算法期望的格式：
        {
            "summary_stats": {...},
            "evaluation_data": [
                {
                    "evaluator_id": "A01",
                    "evaluations": {
                        "B02": 20,  // 總分 (Q1+Q2+Q3+Q4+Q5)
                        "C03": 18,
                        ...
                    }
                }
            ]
        }
        
        Args:
            results: 結果字典
            
        Returns:
            Path: 輸出檔案路徑
        """
        print("\n📊 生成 Vancouver 算法輸入格式...")
        
        # 步驟 1: 彙總每個評分者對每個被評分者的所有問題分數
        evaluator_scores = defaultdict(lambda: defaultdict(int))  # {evaluator_id: {target_id: total_score}}
        question_count = defaultdict(lambda: defaultdict(int))     # {evaluator_id: {target_id: question_count}}
        
        for target_id, data in results['results'].items():
            for eval_data in data['all_evaluations']:
                evaluator_id = eval_data['evaluator_id']
                score = eval_data['score']
                
                # 累加分數
                evaluator_scores[evaluator_id][target_id] += score
                question_count[evaluator_id][target_id] += 1
        
        # 步驟 2: 生成 Vancouver 格式的 evaluation_data
        evaluation_data = []
        for evaluator_id in sorted(evaluator_scores.keys()):
            evaluations_dict = {}
            for target_id, total_score in evaluator_scores[evaluator_id].items():
                evaluations_dict[target_id] = total_score
            
            evaluation_data.append({
                'evaluator_id': evaluator_id,
                'evaluations': evaluations_dict
            })
        
        # 步驟 3: 生成統計摘要
        all_scores = []
        for evaluator_data in evaluation_data:
            all_scores.extend(evaluator_data['evaluations'].values())
        
        summary_stats = {
            'total_evaluators': len(evaluation_data),
            'total_evaluations': sum(len(e['evaluations']) for e in evaluation_data),
            'total_scores': len(all_scores),
            'average_score': sum(all_scores) / len(all_scores) if all_scores else 0,
            'overall_stats': {
                'mean': sum(all_scores) / len(all_scores) if all_scores else 0,
                'min': min(all_scores) if all_scores else 0,
                'max': max(all_scores) if all_scores else 0
            }
        }
        
        # 步驟 4: 組合最終資料
        vancouver_data = {
            'summary_stats': summary_stats,
            'evaluation_data': evaluation_data
        }
        
        # 使用配置生成檔案路徑
        output_path = Path(self.config.get_path('stage5_results', 'vancouver_input', timestamp=datetime.now()))
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(vancouver_data, f, ensure_ascii=False, indent=2)
        
        print(f"✅ Vancouver 格式已匯出: {output_path}")
        print(f"   • 評分者數量: {summary_stats['total_evaluators']}")
        print(f"   • 評分總數: {summary_stats['total_evaluations']}")
        print(f"   • 平均總分: {summary_stats['average_score']:.2f}")
        
        return output_path
    
    def export_to_excel(self, results: Dict, filename: str = None) -> Path:
        """
        匯出結果為 Excel 檔案
        
        Args:
            results: 結果字典
            filename: 檔案名稱（可選）
            
        Returns:
            Path: 輸出檔案路徑
        """
        if not EXCEL_AVAILABLE:
            print("❌ 無法匯出 Excel: openpyxl 未安裝")
            print("   請執行: pip install openpyxl")
            return None
        
        if filename is None:
            # 使用配置生成檔案路徑
            output_path = Path(self.config.get_path('stage5_results', 'evaluation_results_xlsx', timestamp=datetime.now()))
        else:
            output_path = self.output_dir / filename
        
        # 建立工作簿
        wb = openpyxl.Workbook()
        
        # 樣式定義
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 工作表 1: 摘要統計
        ws_summary = wb.active
        ws_summary.title = "摘要統計"
        
        metadata = results['metadata']
        summary = results['summary']
        
        # 寫入摘要資訊
        ws_summary['A1'] = "評分結果摘要"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary.merge_cells('A1:D1')
        
        row = 3
        ws_summary[f'A{row}'] = "生成時間:"
        ws_summary[f'B{row}'] = metadata['generated_at']
        row += 1
        
        ws_summary[f'A{row}'] = "學生總數:"
        ws_summary[f'B{row}'] = metadata['total_students']
        row += 1
        
        ws_summary[f'A{row}'] = "評分總數:"
        ws_summary[f'B{row}'] = metadata['total_evaluations']
        row += 1
        
        ws_summary[f'A{row}'] = "記錄總數:"
        ws_summary[f'B{row}'] = metadata['total_submissions']
        row += 2
        
        # 整體統計
        ws_summary[f'A{row}'] = "整體分數統計"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws_summary[f'A{row}'] = "平均分"
        ws_summary[f'B{row}'] = summary['overall']['mean']
        row += 1
        
        ws_summary[f'A{row}'] = "最低分"
        ws_summary[f'B{row}'] = summary['overall']['min']
        row += 1
        
        ws_summary[f'A{row}'] = "最高分"
        ws_summary[f'B{row}'] = summary['overall']['max']
        row += 2
        
        # 各題統計
        ws_summary[f'A{row}'] = "各題平均分"
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # 表頭
        ws_summary[f'A{row}'] = "問題"
        ws_summary[f'B{row}'] = "平均分"
        ws_summary[f'C{row}'] = "最低分"
        ws_summary[f'D{row}'] = "最高分"
        for col in ['A', 'B', 'C', 'D']:
            ws_summary[f'{col}{row}'].fill = header_fill
            ws_summary[f'{col}{row}'].font = header_font
            ws_summary[f'{col}{row}'].alignment = center_alignment
        row += 1
        
        for q_id in sorted(summary['by_question'].keys()):
            stats = summary['by_question'][q_id]
            ws_summary[f'A{row}'] = q_id
            ws_summary[f'B{row}'] = stats['mean']
            ws_summary[f'C{row}'] = stats['min']
            ws_summary[f'D{row}'] = stats['max']
            row += 1
        
        # 工作表 2: 學生詳細結果
        ws_details = wb.create_sheet("學生詳細結果")
        
        headers = ["學生ID", "姓名", "評分數", "總記錄數", "平均分", "最低分", "最高分", 
                   "Q1平均", "Q2平均", "Q3平均", "Q4平均", "Q5平均"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        row = 2
        for target_id in sorted(results['results'].keys()):
            data = results['results'][target_id]
            
            ws_details.cell(row=row, column=1, value=data['student_id'])
            ws_details.cell(row=row, column=2, value=data['student_name'])
            ws_details.cell(row=row, column=3, value=data['total_evaluations'])
            ws_details.cell(row=row, column=4, value=data['total_scores'])
            ws_details.cell(row=row, column=5, value=data['overall_mean'])
            ws_details.cell(row=row, column=6, value=data['overall_min'])
            ws_details.cell(row=row, column=7, value=data['overall_max'])
            
            # 各題平均
            for q_idx, q_id in enumerate(['Q1', 'Q2', 'Q3', 'Q4', 'Q5'], 8):
                if q_id in data['question_stats']:
                    ws_details.cell(row=row, column=q_idx, 
                                  value=data['question_stats'][q_id]['mean'])
            
            # 添加邊框
            for col_idx in range(1, len(headers) + 1):
                ws_details.cell(row=row, column=col_idx).border = border
                ws_details.cell(row=row, column=col_idx).alignment = center_alignment
            
            row += 1
        
        # 工作表 3: 所有評分記錄
        ws_all = wb.create_sheet("所有評分記錄")
        
        headers = ["評分者ID", "被評分者ID", "被評分者姓名", "問題ID", "分數", "評論", "提交時間"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_all.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        row = 2
        for target_id in sorted(results['results'].keys()):
            data = results['results'][target_id]
            
            for eval_data in sorted(data['all_evaluations'], 
                                   key=lambda x: (x['evaluator_id'], x['question_id'])):
                ws_all.cell(row=row, column=1, value=eval_data['evaluator_id'])
                ws_all.cell(row=row, column=2, value=target_id)
                ws_all.cell(row=row, column=3, value=data['student_name'])
                ws_all.cell(row=row, column=4, value=eval_data['question_id'])
                ws_all.cell(row=row, column=5, value=eval_data['score'])
                ws_all.cell(row=row, column=6, value=eval_data['comment'])
                ws_all.cell(row=row, column=7, value=eval_data['submitted_at'])
                
                for col_idx in range(1, len(headers) + 1):
                    ws_all.cell(row=row, column=col_idx).border = border
                
                row += 1
        
        # 調整列寬
        for ws in [ws_summary, ws_details, ws_all]:
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
        
        # 儲存
        wb.save(output_path)
        
        print(f"✅ Excel 檔案已匯出: {output_path}")
        print(f"   檔案大小: {output_path.stat().st_size / 1024:.2f} KB")
        print(f"   工作表: 摘要統計, 學生詳細結果, 所有評分記錄")
        
        return output_path
    
    def print_summary(self, results: Dict):
        """列印結果摘要"""
        print("\n" + "=" * 80)
        print("📊 評分結果摘要")
        print("=" * 80)
        
        metadata = results['metadata']
        summary = results['summary']
        
        print(f"\n基本統計:")
        print(f"  學生總數: {metadata['total_students']}")
        print(f"  評分總數: {metadata['total_evaluations']}")
        print(f"  記錄總數: {metadata['total_submissions']}")
        
        print(f"\n整體分數:")
        print(f"  平均分: {summary['overall']['mean']}")
        print(f"  最低分: {summary['overall']['min']}")
        print(f"  最高分: {summary['overall']['max']}")
        
        print(f"\n各題平均分:")
        for q_id, stats in sorted(summary['by_question'].items()):
            print(f"  {q_id}: {stats['mean']} (範圍: {stats['min']}-{stats['max']})")
        
        print("\n" + "=" * 80)


def main():
    """主函數"""
    collector = ResultCollectorWeb()
    
    # 收集結果
    results = collector.collect_results()
    
    # 列印摘要
    collector.print_summary(results)
    
    # 匯出 JSON
    json_path = collector.export_to_json(results)
    
    # 匯出 Excel
    excel_path = collector.export_to_excel(results)
    
    # 匯出 Vancouver 格式
    vancouver_path = collector.export_for_vancouver(results)
    
    print("\n✅ 結果收集完成！")
    print(f"\n輸出檔案:")
    print(f"  1. 完整結果 (JSON): {json_path}")
    if excel_path:
        print(f"  2. 完整結果 (Excel): {excel_path}")
    print(f"  3. Vancouver 格式: {vancouver_path}")


if __name__ == '__main__':
    main()
